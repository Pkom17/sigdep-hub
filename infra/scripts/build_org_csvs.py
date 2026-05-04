#!/usr/bin/env python3
"""
Read organization.xlsx (Regions / Districts / Sites sheets) and emit three CSVs
ready to be loaded by Liquibase loadData into core.regions / core.districts /
core.sites.

Usage:
    python3 build_org_csvs.py [path/to/organization.xlsx] [output_dir]

Defaults:
    input  : ../../organization.xlsx (relative to this script, i.e. repo root)
    output : ../ingestion-api/src/main/resources/db/data/

The generated CSVs use header columns matching the Liquibase loadData mapping.
Foreign keys are resolved by looking up parent rows by their natural code.
"""

import csv
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    here = Path(__file__).resolve().parent
    default_xlsx = here.parent.parent.parent / "organization.xlsx"
    default_out = (
        here.parent.parent
        / "ingestion-api"
        / "src"
        / "main"
        / "resources"
        / "db"
        / "data"
    )

    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_xlsx
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else default_out

    if not xlsx_path.exists():
        print(f"ERROR: input file not found: {xlsx_path}", file=sys.stderr)
        return 1
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Reading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # --- Regions ----------------------------------------------------------
    ws = wb["Regions"]
    regions_path = out_dir / "regions.csv"
    region_code_by_name: dict[str, str] = {}
    n_regions = 0
    with regions_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["code", "name", "source_uuid"])
        for row in ws.iter_rows(min_row=2, values_only=True):
            rid, name, uuid = row
            if rid is None or not name:
                continue
            code = str(rid)
            w.writerow([code, name.strip(), uuid or ""])
            region_code_by_name[name.strip()] = code
            n_regions += 1
    print(f"  wrote {n_regions} rows -> {regions_path}")

    # --- Districts --------------------------------------------------------
    ws = wb["Districts"]
    districts_path = out_dir / "districts.csv"
    district_code_by_pair: dict[tuple[str, str], str] = {}
    n_districts = 0
    n_district_orphans = 0
    with districts_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["code", "region_code", "name", "source_uuid"])
        for row in ws.iter_rows(min_row=2, values_only=True):
            did, region_name, district_name, uuid = row
            if did is None or not district_name:
                continue
            region_name = (region_name or "").strip()
            district_name = district_name.strip()
            region_code = region_code_by_name.get(region_name)
            if region_code is None:
                n_district_orphans += 1
                continue
            code = str(did)
            w.writerow([code, region_code, district_name, uuid or ""])
            district_code_by_pair[(region_name, district_name)] = code
            n_districts += 1
    print(f"  wrote {n_districts} rows -> {districts_path}")
    if n_district_orphans:
        print(f"  WARN: {n_district_orphans} district rows skipped (region not found)")

    # --- Sites ------------------------------------------------------------
    ws = wb["Sites"]
    sites_path = out_dir / "sites.csv"
    n_sites = 0
    n_site_orphans = 0
    with sites_path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["code", "name", "district_code", "facility_type", "source_uuid"])
        for row in ws.iter_rows(min_row=2, values_only=True):
            region_name, district_name, ident, name, ftype, uuid = row
            if not ident or not name:
                continue
            pair = ((region_name or "").strip(), (district_name or "").strip())
            district_code = district_code_by_pair.get(pair)
            if district_code is None:
                n_site_orphans += 1
                continue
            w.writerow(
                [
                    str(ident).strip(),
                    name.strip(),
                    district_code,
                    (ftype or "").strip(),
                    uuid or "",
                ]
            )
            n_sites += 1
    print(f"  wrote {n_sites} rows -> {sites_path}")
    if n_site_orphans:
        print(f"  WARN: {n_site_orphans} site rows skipped (district not found)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
